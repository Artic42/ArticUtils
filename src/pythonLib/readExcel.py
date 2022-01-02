import openpyxl

def readExcelFile (path, sheetName):
    file = openpyxl.load_workbook (path, read_only=True)
    sheets = file.sheetnames
    return file.get_sheet_by_name (sheetName)

def getValueCell (worksheet, row, col):
    cell = worksheet.cell (row, col)
    return cell.value

def getRowValues (worksheet, start, end):
    rows = worksheet.iter_rows (min_row=start, max_row=end, values_only=False)
    result = []
    for row in rows:
        rowValues = []
        for cell in row:
            rowValues.append (cell.value)
        result.append (rowValues)
    return result






